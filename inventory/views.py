from django.shortcuts import render, redirect
from .models import Prenda
from django.http import HttpResponse
from .forms import PrendaForm
import unicodedata, random
from inventory.models import Subcategoria, Categoria
from django.contrib.auth.decorators import login_required
from django.shortcuts import get_object_or_404, redirect
from django.db.models import Sum, F, ExpressionWrapper, DecimalField
from django.utils import timezone
from django.db.models.functions import TruncDate
from django.contrib import messages
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from django.utils.timezone import localtime


def normalizar(texto):
    """Normaliza el texto para búsquedas insensibles a tildes y mayúsculas"""
    texto = str(texto).lower()
    return ''.join(c for c in unicodedata.normalize('NFD', texto) if unicodedata.category(c) != 'Mn')


def lista_prendas(request):
    query = request.GET.get('q', '').strip()
    prendas = Prenda.objects.filter(disponible=True)

    # Obtener prendas recomendadas de forma aleatoria
    prendas_recomendadas = Prenda.objects.filter(disponible=True).order_by('?')[:5]

    if query:
        query_norm = normalizar(query)

        if len(query) == 1:
            prendas = prendas.filter(talla__iexact=query)
        else:
            # Primero buscamos coincidencias en grupos
            grupos_busqueda = {
                'mujer': 'Mujer',
                'hombre': 'Hombre',
                'niño': 'Niño',
                'nino': 'Niño',
                'bebe': 'Bebé',
                'bebé': 'Bebé'
            }
            
            grupo_encontrado = grupos_busqueda.get(query_norm)
            if grupo_encontrado:
                prendas = prendas.filter(subcategoria__categoria__grupo__nombre=grupo_encontrado)
            else:
                # Buscamos coincidencias en categorías y subcategorías
                categorias_match = Categoria.objects.filter(nombre__icontains=query)
                subcategorias_match = Subcategoria.objects.filter(nombre__icontains=query)
                
                # Creamos una lista para almacenar todas las prendas que coinciden
                prendas_match = []
                
                # Agregamos prendas que coinciden por categoría
                for categoria in categorias_match:
                    prendas_match.extend(list(prendas.filter(subcategoria__categoria=categoria)))
                
                # Agregamos prendas que coinciden por subcategoría
                for subcategoria in subcategorias_match:
                    prendas_match.extend(list(prendas.filter(subcategoria=subcategoria)))
                
                # Agregamos prendas que coinciden por nombre o descripción
                prendas_match.extend([
                    prenda for prenda in prendas if
                    query_norm in normalizar(prenda.nombre) or
                    query_norm in normalizar(prenda.descripcion)
                ])
                
                # Eliminamos duplicados manteniendo el orden
                prendas = list(dict.fromkeys(prendas_match))

    return render(request, 'lista_prendas.html', {
        'prendas': prendas,
        'prendas_recomendadas': prendas_recomendadas
    })

@login_required
def marcar_como_disponible(request, prenda_id):
    prenda = get_object_or_404(Prenda, id=prenda_id)
    prenda.disponible = True
    prenda.save()
    return redirect('ver_ventas')

def home(request):
    prendas_recientes = Prenda.objects.filter(disponible=True).order_by('-fecha_agregado')[:10]
    todas = list(Prenda.objects.filter(disponible=True))
    prendas_random = random.sample(todas, min(len(todas), 3))

    return render(request, 'home.html', {
        'prendas_recientes': prendas_recientes,
        'prendas_random': prendas_random
    })

def catalogo_recientes(request):
    prendas_recientes = Prenda.objects.filter(disponible=True).order_by('-fecha_agregado')[:10]
    return render(request, 'recientes.html', {'prendas_recientes': prendas_recientes})

def contacto(request):
    return render(request, 'contacto.html')

@login_required
def ver_ventas(request):
    prendas_vendidas = Prenda.objects.filter(disponible=False)

    desde = request.GET.get('desde')
    hasta = request.GET.get('hasta')
    categoria_id = request.GET.get('categoria')
    subcategoria_id = request.GET.get('subcategoria')

    if desde:
        prendas_vendidas = prendas_vendidas.filter(fecha_venta__date__gte=desde)
    if hasta:
        prendas_vendidas = prendas_vendidas.filter(fecha_venta__date__lte=hasta)
    if subcategoria_id:
        prendas_vendidas = prendas_vendidas.filter(subcategoria_id=subcategoria_id)

    total = prendas_vendidas.aggregate(Sum('precio'))['precio__sum'] or 0

    categorias = Categoria.objects.all()
    subcategorias = Subcategoria.objects.select_related('categoria')

    return render(request, 'ventas.html', {
        'prendas': prendas_vendidas,
        'categorias': categorias,
        'subcategorias': subcategorias,
        'total_ventas': total,
    })

@login_required
def marcar_como_vendida(request, prenda_id):
    prenda = get_object_or_404(Prenda, id=prenda_id)
    prenda.disponible = False
    prenda.fecha_venta = timezone.now()
    prenda.save()

    # Aquí puedes crear el registro de venta si tienes modelo Venta
    return redirect('ver_inventario')
@login_required
def admin_dashboard(request):
    prendas_disponibles = Prenda.objects.filter(disponible=True).count()
    prendas_vendidas = Prenda.objects.filter(disponible=False).count()

    prendas_vendidas_qs = Prenda.objects.filter(disponible=False)

    total_ventas = prendas_vendidas_qs.aggregate(total=Sum('precio'))['total'] or 0
    total_inversion = prendas_vendidas_qs.aggregate(total=Sum('precio_proveedor'))['total'] or 0

    ganancias_zero_waste = prendas_vendidas_qs.aggregate(
        total=Sum(
            ExpressionWrapper(
                F('precio') - F('precio_proveedor'),
                output_field=DecimalField()
            )
        )
    )['total'] or 0

    return render(request, 'admin_dashboard.html', {
        'prendas_disponibles': prendas_disponibles,
        'prendas_vendidas': prendas_vendidas,
        'total_ventas': total_ventas,
        'total_inversion': total_inversion,
        'ganancias_zero_waste': ganancias_zero_waste,
    })

@login_required
def agregar_prenda(request):
    if request.method == 'POST':
        form = PrendaForm(request.POST, request.FILES)
        if form.is_valid():
            prenda = form.save(commit=False)
            prenda.disponible = True  # Siempre disponible al registrarla

            # 🔥 Asignar subcategoría seleccionada manualmente
            subcategoria_id = request.POST.get('subcategoria')
            if subcategoria_id:
                try:
                    prenda.subcategoria = Subcategoria.objects.get(id=subcategoria_id)
                except Subcategoria.DoesNotExist:
                    pass  # o puedes agregar un mensaje de error si quieres

            prenda.save()
            return redirect('admin_dashboard')
    else:
        form = PrendaForm()

    subcategorias = Subcategoria.objects.select_related('categoria')

    return render(request, 'agregar_prenda.html', {
        'form': form,
        'subcategorias': subcategorias
    })

@login_required
def editar_prenda(request, prenda_id):
    prenda = get_object_or_404(Prenda, id=prenda_id)

    if request.method == 'POST':
        form = PrendaForm(request.POST, request.FILES, instance=prenda)
        if form.is_valid():
            prenda = form.save(commit=False)

            # ✅ Asignar subcategoría solo si cambia
            subcategoria_id = request.POST.get('subcategoria')
            if subcategoria_id:
                try:
                    subcat = Subcategoria.objects.get(id=subcategoria_id)
                    prenda.subcategoria = subcat
                except Subcategoria.DoesNotExist:
                    pass

            prenda.save()
            return redirect('ver_inventario')
    else:
        form = PrendaForm(instance=prenda)

    subcategorias = Subcategoria.objects.select_related('categoria')

    return render(request, 'editar_prenda.html', {
        'form': form,
        'subcategorias': subcategorias,
        'prenda': prenda,
    })



@login_required
def ver_inventario(request):
    query_codigo = request.GET.get('buscar_codigo')
    prendas = Prenda.objects.filter(disponible=True).order_by('-fecha_agregado')
    if query_codigo:
        prendas = prendas.filter(codigo__icontains=query_codigo)
    return render(request, 'ver_inventario.html', {'prendas': prendas})

    return render(request, 'ver_inventario.html', {'prendas': prendas})
def exportar_ventas_excel(request):
    prendas_vendidas = Prenda.objects.filter(disponible=False)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ventas Zero Waste"

    # Encabezados
    columnas = [
        'Código', 'Nombre', 'Talla', 'Precio Venta', 'Precio Proveedor', 'Ganancia',
        'Categoría', 'Subcategoría', 'Nombre Proveedor', 'Fecha de Venta'
    ]

    ws.append(columnas)

    # Estilo de encabezados
    for col_num, col_title in enumerate(columnas, 1):
        col_letter = get_column_letter(col_num)
        ws[f'{col_letter}1'].font = Font(bold=True)

    # Filas de datos
    for prenda in prendas_vendidas:
        categoria = prenda.subcategoria.categoria.nombre if prenda.subcategoria and prenda.subcategoria.categoria else ''
        subcategoria = prenda.subcategoria.nombre if prenda.subcategoria else ''
        ganancia = prenda.precio - prenda.precio_proveedor

        ws.append([
            prenda.codigo,
            prenda.nombre,
            prenda.talla,
            float(prenda.precio),
            float(prenda.precio_proveedor),
            float(ganancia),
            categoria,
            subcategoria,
            prenda.nombre_proveedor,
            localtime(prenda.fecha_venta).strftime('%d/%m/%Y') if prenda.fecha_venta else ''

        ])

    # Ajuste de anchos
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length + 2

    # Preparar la respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=ventas_zerowaste.xlsx'
    wb.save(response)
    return response

@login_required
def ver_estadisticas(request):
    prendas_vendidas = Prenda.objects.filter(disponible=False)

    ventas_fecha = prendas_vendidas.annotate(
        fecha=TruncDate('fecha_venta')
    ).values('fecha').annotate(total=Sum('precio')).order_by('fecha')

    total_ventas = prendas_vendidas.aggregate(total=Sum('precio'))['total'] or 0
    ganancia_proveedores = prendas_vendidas.aggregate(total=Sum('precio_proveedor'))['total'] or 0
    ganancia_zero_waste = prendas_vendidas.aggregate(
        total=Sum(ExpressionWrapper(F('precio') - F('precio_proveedor'), output_field=DecimalField()))
    )['total'] or 0

    return render(request, 'estadisticas.html', {
        'ventas_fecha': ventas_fecha,
        'total_ventas': total_ventas,
        'ganancia_proveedores': ganancia_proveedores,
        'ganancia_zero_waste': ganancia_zero_waste,
    })


@login_required
def ver_estadisticas_filtradas(request):
    prendas_vendidas = Prenda.objects.filter(disponible=False)

    # Filtros
    desde = request.GET.get('desde')
    hasta = request.GET.get('hasta')
    grupo = request.GET.get('grupo')
    categoria = request.GET.get('categoria')

    if desde:
        prendas_vendidas = prendas_vendidas.filter(fecha_venta__date__gte=desde)
    if hasta:
        prendas_vendidas = prendas_vendidas.filter(fecha_venta__date__lte=hasta)
    if grupo:
        prendas_vendidas = prendas_vendidas.filter(subcategoria__categoria__grupo__nombre=grupo)
    if categoria:
        prendas_vendidas = prendas_vendidas.filter(subcategoria__categoria__nombre=categoria)
    # Filtrar categorías por grupo actual
    if grupo:
        categorias_filtradas = Categoria.objects.filter(grupo__nombre=grupo).order_by('nombre').distinct('nombre')
    else:
        categorias_filtradas = Categoria.objects.none()

    # Gráficas
    ventas_subcat = prendas_vendidas.values('subcategoria__nombre').annotate(total=Sum('precio')).order_by('-total')
    ventas_categoria = prendas_vendidas.values('subcategoria__categoria__nombre').annotate(total=Sum('precio')).order_by('-total')
    ventas_fecha = prendas_vendidas.annotate(
        fecha=TruncDate('fecha_venta')
    ).values('fecha').annotate(total=Sum('precio')).order_by('fecha')

    # Totales
    total_ventas = prendas_vendidas.aggregate(Sum('precio'))['precio__sum'] or 0
    ganancia_proveedores = prendas_vendidas.aggregate(Sum('precio_proveedor'))['precio_proveedor__sum'] or 0
    ganancia_zero_waste = prendas_vendidas.aggregate(
        total=Sum(ExpressionWrapper(F('precio') - F('precio_proveedor'), output_field=DecimalField()))
    )['total'] or 0

    categorias = Categoria.objects.all()

    return render(request, 'estadisticas_filtradas.html', {
        'ventas_subcat': ventas_subcat,
        'ventas_categoria': ventas_categoria,
        'ventas_fecha': ventas_fecha,
        'total_ventas': total_ventas,
        'ganancia_proveedores': ganancia_proveedores,
        'ganancia_zero_waste': ganancia_zero_waste,
        'desde': desde,
        'hasta': hasta,
        'grupo': grupo,
        'categoria': categoria,
        'categorias': categorias_filtradas,

    })



@login_required
def eliminar_prenda(request, prenda_id):
    prenda = get_object_or_404(Prenda, id=prenda_id)
    prenda.delete()
    messages.success(request, "✅ La prenda fue eliminada exitosamente.")
    return redirect('ver_inventario')

# catalogoooo
def catalogo_bebebodies(request):
    prendas = Prenda.objects.filter(
        subcategoria__nombre='Bodies',
        subcategoria__categoria__grupo__nombre='Bebé',
        disponible=True
    )
    return render(request, 'categorias/Bebe/BebeBodies.html', {'prendas': prendas})


def catalogo_bebeconjuntos(request):
    prendas = Prenda.objects.filter(
        subcategoria__nombre='Conjuntos',
        subcategoria__categoria__grupo__nombre='Bebé',
        disponible=True
    )
    return render(request, 'categorias/Bebe/BebeConjuntos.html', {'prendas': prendas})


def catalogo_bebepijamas(request):
    prendas = Prenda.objects.filter(
        subcategoria__nombre='Pijamas',
        subcategoria__categoria__grupo__nombre='Bebé',
        disponible=True
    )
    return render(request, 'categorias/Bebe/BebePijamas.html', {'prendas': prendas})


def catalogo_niñocamisas(request):
    prendas = Prenda.objects.filter(
        subcategoria__nombre='Camisas',
        subcategoria__categoria__grupo__nombre='Niño',
        disponible=True
    )
    return render(request, 'categorias/Niño/NiñoCamisas.html', {'prendas': prendas})


def catalogo_niñochaquetas(request):
    prendas = Prenda.objects.filter(
        subcategoria__nombre='Chaquetas',
        subcategoria__categoria__grupo__nombre='Niño',
        disponible=True
    )
    return render(request, 'categorias/Niño/NiñoChaquetas.html', {'prendas': prendas})


def catalogo_niñopantalones(request):
    prendas = Prenda.objects.filter(
        subcategoria__nombre='Pantalones',
        subcategoria__categoria__grupo__nombre='Niño',
        disponible=True
    )
    return render(request, 'categorias/Niño/NiñoPantalones.html', {'prendas': prendas})


def catalogo_niñozapatos(request):
    prendas = Prenda.objects.filter(
        subcategoria__nombre='Zapatos',
        subcategoria__categoria__grupo__nombre='Niño',
        disponible=True
    )
    return render(request, 'categorias/Niño/NiñoZapatos.html', {'prendas': prendas})


def catalogo_hombrecamisas(request):
    prendas = Prenda.objects.filter(
        subcategoria__nombre='Camisas',
        subcategoria__categoria__grupo__nombre='Hombre',
        disponible=True
    )
    return render(request, 'categorias/Hombre/HombreCamisas.html', {'prendas': prendas})


def catalogo_hombrechaquetas(request):
    prendas = Prenda.objects.filter(
        subcategoria__nombre='Chaquetas',
        subcategoria__categoria__grupo__nombre='Hombre',
        disponible=True
    )
    return render(request, 'categorias/Hombre/HombreChaquetas.html', {'prendas': prendas})


def catalogo_hombrepantalones(request):
    prendas = Prenda.objects.filter(
        subcategoria__nombre='Pantalones',
        subcategoria__categoria__grupo__nombre='Hombre',
        disponible=True
    )
    return render(request, 'categorias/Hombre/HombrePantalones.html', {'prendas': prendas})


def catalogo_hombrezapatos(request):
    prendas = Prenda.objects.filter(
        subcategoria__nombre='Zapatos',
        subcategoria__categoria__grupo__nombre='Hombre',
        disponible=True
    )
    return render(request, 'categorias/Hombre/HombreZapatos.html', {'prendas': prendas})


def catalogo_hombredeportivos(request):
    prendas = Prenda.objects.filter(
        subcategoria__nombre='Deportivos',
        subcategoria__categoria__grupo__nombre='Hombre',
        disponible=True
    )
    return render(request, 'categorias/Hombre/HombreDeportivos.html', {'prendas': prendas})


def catalogo_hombreblazers(request):
    prendas = Prenda.objects.filter(
        subcategoria__nombre='Blazers',
        subcategoria__categoria__grupo__nombre='Hombre',
        disponible=True
    )
    return render(request, 'categorias/Hombre/HombreBlazers.html', {'prendas': prendas})


def catalogo_mujerbolsos(request):
    prendas = Prenda.objects.filter(
        subcategoria__nombre='Bolsos',
        subcategoria__categoria__grupo__nombre='Mujer',
        disponible=True
    )
    return render(request, 'categorias/Mujer/Accesorios(Mujer)/MujerBolsos.html', {'prendas': prendas})


def catalogo_mujerzapatos(request):
    prendas = Prenda.objects.filter(
        subcategoria__nombre='Zapatos',
        subcategoria__categoria__grupo__nombre='Mujer',
        disponible=True
    )
    return render(request, 'categorias/Mujer/Accesorios(Mujer)/MujerZapatos.html', {'prendas': prendas})


def catalogo_mujersandalias(request):
    prendas = Prenda.objects.filter(
        subcategoria__nombre='Sandalias',
        subcategoria__categoria__grupo__nombre='Mujer',
        disponible=True
    )
    return render(request, 'categorias/Mujer/Accesorios(Mujer)/MujerSandalias.html', {'prendas': prendas})


def catalogo_mujeraccesorios(request):
    prendas = Prenda.objects.filter(
        subcategoria__nombre='Accesorios',
        subcategoria__categoria__grupo__nombre='Mujer',
        disponible=True
    )
    return render(request, 'categorias/Mujer/Accesorios(Mujer)/MujerAccesorios.html', {'prendas': prendas})


def catalogo_mujerropainterior(request):
    prendas = Prenda.objects.filter(
        subcategoria__nombre='Ropa interior',
        subcategoria__categoria__grupo__nombre='Mujer',
        disponible=True
    )
    return render(request, 'categorias/Mujer/Especiales(Mujer)/MujerRopaInterior.html', {'prendas': prendas})


def catalogo_mujerpijamas(request):
    prendas = Prenda.objects.filter(
        subcategoria__nombre='Pijamas',
        subcategoria__categoria__grupo__nombre='Mujer',
        disponible=True
    )
    return render(request, 'categorias/Mujer/Especiales(Mujer)/MujerPijamas.html', {'prendas': prendas})


def catalogo_mujerdeportivos(request):
    prendas = Prenda.objects.filter(
        subcategoria__nombre='Deportivos',
        subcategoria__categoria__grupo__nombre='Mujer',
        disponible=True
    )
    return render(request, 'categorias/Mujer/Especiales(Mujer)/MujerDeportivos.html', {'prendas': prendas})


def catalogo_mujerconjuntos(request):
    prendas = Prenda.objects.filter(
        subcategoria__nombre='Conjuntos',
        subcategoria__categoria__grupo__nombre='Mujer',
        disponible=True
    )
    return render(request, 'categorias/Mujer/Especiales(Mujer)/MujerConjuntos.html', {'prendas': prendas})


def catalogo_mujertrajesdebaño(request):
    prendas = Prenda.objects.filter(
        subcategoria__nombre='Trajes de baño',
        subcategoria__categoria__grupo__nombre='Mujer',
        disponible=True
    )
    return render(request, 'categorias/Mujer/Especiales(Mujer)/MujerTrajesDeBaño.html', {'prendas': prendas})


def catalogo_mujerbodies(request):
    prendas = Prenda.objects.filter(
        subcategoria__nombre='Bodies',
        subcategoria__categoria__grupo__nombre='Mujer',
        disponible=True
    )
    return render(request, 'categorias/Mujer/Inferior(Mujer)/MujerBodies.html', {'prendas': prendas})


def catalogo_mujershorts(request):
    prendas = Prenda.objects.filter(
        subcategoria__nombre='Shorts',
        subcategoria__categoria__grupo__nombre='Mujer',
        disponible=True
    )
    return render(request, 'categorias/Mujer/Inferior(Mujer)/MujerShorts.html', {'prendas': prendas})


def catalogo_mujerfaldas(request):
    prendas = Prenda.objects.filter(
        subcategoria__nombre='Faldas',
        subcategoria__categoria__grupo__nombre='Mujer',
        disponible=True
    )
    return render(request, 'categorias/Mujer/Inferior(Mujer)/MujerFaldas.html', {'prendas': prendas})


def catalogo_mujerjeans(request):
    prendas = Prenda.objects.filter(
        subcategoria__nombre='Jeans',
        subcategoria__categoria__grupo__nombre='Mujer',
        disponible=True
    )
    return render(request, 'categorias/Mujer/Inferior(Mujer)/MujerJeans.html', {'prendas': prendas})


def catalogo_mujerpantalones(request):
    prendas = Prenda.objects.filter(
        subcategoria__nombre='Pantalones',
        subcategoria__categoria__grupo__nombre='Mujer',
        disponible=True
    )
    return render(request, 'categorias/Mujer/Inferior(Mujer)/MujerPantalones.html', {'prendas': prendas})


def catalogo_mujervestidos(request):
    prendas = Prenda.objects.filter(
        subcategoria__nombre='Vestidos',
        subcategoria__categoria__grupo__nombre='Mujer',
        disponible=True
    )
    return render(request, 'categorias/Mujer/Ropa(Mujer)/MujerVestidos.html', {'prendas': prendas})


def catalogo_mujertops(request):
    prendas = Prenda.objects.filter(
        subcategoria__nombre='Tops',
        subcategoria__categoria__grupo__nombre='Mujer',
        disponible=True
    )
    return render(request, 'categorias/Mujer/Ropa(Mujer)/MujerTops.html', {'prendas': prendas})


def catalogo_mujerenterizos(request):
    prendas = Prenda.objects.filter(
        subcategoria__nombre='Enterizos',
        subcategoria__categoria__grupo__nombre='Mujer',
        disponible=True
    )
    return render(request, 'categorias/Mujer/Ropa(Mujer)/MujerEnterizos.html', {'prendas': prendas})


def catalogo_mujercamisetas(request):
    prendas = Prenda.objects.filter(
        subcategoria__nombre='Camisetas',
        subcategoria__categoria__grupo__nombre='Mujer',
        disponible=True
    )
    return render(request, 'categorias/Mujer/Ropa(Mujer)/MujerCamisetas.html', {'prendas': prendas})


def catalogo_mujercamisas(request):
    prendas = Prenda.objects.filter(
        subcategoria__nombre='Camisas',
        subcategoria__categoria__grupo__nombre='Mujer',
        disponible=True
    )
    return render(request, 'categorias/Mujer/Ropa(Mujer)/MujerCamisas.html', {'prendas': prendas})


def catalogo_mujerchaquetas(request):
    prendas = Prenda.objects.filter(
        subcategoria__nombre='Chaquetas',
        subcategoria__categoria__grupo__nombre='Mujer',
        disponible=True
    )
    return render(request, 'categorias/Mujer/Ropa(Mujer)/MujerChaquetas.html', {'prendas': prendas})


def catalogo_mujerblusas(request):
    prendas = Prenda.objects.filter(
        subcategoria__nombre='Blusas',
        subcategoria__categoria__grupo__nombre='Mujer',
        disponible=True
    )
    return render(request, 'categorias/Mujer/Ropa(Mujer)/MujerBlusas.html', {'prendas': prendas})

def about(request):
    return render(request, 'about.html')
